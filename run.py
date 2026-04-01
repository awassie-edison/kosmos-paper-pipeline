"""Pipeline orchestrator with checkpoint/resume and CLI interface."""

import argparse
import json
import logging
import sys
from datetime import date
from pathlib import Path

from . import config
from .status import StatusTracker

log = logging.getLogger("kosmos_pipeline")

# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def _save(data, path: Path):
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _load(path: Path):
    with open(path) as f:
        return json.load(f)


def _checkpoint_exists(path: Path, resume: bool) -> bool:
    return resume and path.exists()


# ---------------------------------------------------------------------------
# Exclude-DOI loading
# ---------------------------------------------------------------------------

def _load_exclude_dois(path: str | None) -> set[str]:
    """Load DOIs to exclude from a text file (one per line) or xlsx."""
    if not path:
        return set()
    p = Path(path)
    if not p.exists():
        log.warning("Exclude file not found: %s", path)
        return set()

    if p.suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True, data_only=True)
            dois = set()
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and cell.startswith("10."):
                            dois.add(cell.lower().strip())
            log.info("Loaded %d DOIs to exclude from %s", len(dois), path)
            return dois
        except ImportError:
            log.error("openpyxl required to read xlsx files. pip install openpyxl")
            return set()
    else:
        # Plain text, one DOI per line
        dois = set()
        for line in p.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                dois.add(line.lower())
        log.info("Loaded %d DOIs to exclude from %s", len(dois), path)
        return dois


# ---------------------------------------------------------------------------
# Pipeline steps
# ---------------------------------------------------------------------------

def run_pipeline(
    output_dir: str,
    resume: bool = True,
    max_candidates: int = config.MAX_CANDIDATES_TO_EVALUATE,
    max_papers: int = config.MAX_PAPERS_OUTPUT,
    model: str = config.DEFAULT_MODEL,
    exclude_dois_file: str | None = None,
    skip_download: bool = False,
    skip_upload: bool = False,
    review: bool = False,
    stage: str = "DEV",
    kosmos_opt_dir: str | None = None,
):
    """Run the full paper search and curation pipeline."""
    run_dir = Path(output_dir)
    intermediate = run_dir / "intermediate"
    intermediate.mkdir(parents=True, exist_ok=True)

    status = StatusTracker(run_dir / "status.json")

    start_date, end_date = config.date_window()
    log.info("Pipeline start: window %s to %s, output → %s", start_date, end_date, run_dir)

    # Track results across steps for the manifest
    pdf_results: dict = {}
    upload_results: dict = {}

    # Load exclude DOIs
    exclude_dois = _load_exclude_dois(exclude_dois_file)

    # ------------------------------------------------------------------
    # STEP 1: Search
    # ------------------------------------------------------------------
    step1_path = intermediate / "step1_search_results.json"
    if _checkpoint_exists(step1_path, resume):
        log.info("Step 1: loading from checkpoint")
        all_papers = _load(step1_path)
    else:
        status.update("step1_search", "Launching parallel search queries...")
        from .search import run_search
        all_papers = run_search(start_date, end_date)
        _save(all_papers, step1_path)
        status.update("step1_search", "Search complete", {
            "total_raw_results": len(all_papers),
            "unique_dois": len({p["doi"] for p in all_papers}),
        })

    # ------------------------------------------------------------------
    # STEP 2: Filter
    # ------------------------------------------------------------------
    step2_path = intermediate / "step2_filtered.json"
    if _checkpoint_exists(step2_path, resume):
        log.info("Step 2: loading from checkpoint")
        filter_result = _load(step2_path)
        candidates = filter_result["candidates"]
    else:
        status.update("step2_filter", "Filtering by journal tier and article type...")
        from .filtering import filter_papers
        filter_result = filter_papers(all_papers, exclude_dois=exclude_dois)
        candidates = filter_result["candidates"]
        _save(filter_result, step2_path)
        status.update("step2_filter", "Filtering complete", {
            "tier1": filter_result["tier1"],
            "tier2": filter_result["tier2"],
            "excluded_tier0": filter_result["excluded_tier0"],
            "excluded_type": filter_result["excluded_type"],
            "candidates_remaining": len(candidates),
        })

    # Cap candidates for evaluation
    candidates_to_eval = candidates[:max_candidates]
    log.info(
        "Evaluating top %d of %d candidates",
        len(candidates_to_eval),
        len(candidates),
    )

    # ------------------------------------------------------------------
    # STEP 3: Evaluate via Claude API
    # ------------------------------------------------------------------
    step3_path = intermediate / "step3_evaluated.json"
    if _checkpoint_exists(step3_path, resume):
        log.info("Step 3: loading from checkpoint")
        evaluated = _load(step3_path)
    else:
        status.update("step3_evaluate", f"Evaluating {len(candidates_to_eval)} candidates...", {
            "total_candidates": len(candidates_to_eval),
        })
        from .evaluate import evaluate_papers
        evaluated = evaluate_papers(candidates_to_eval, model=model)
        _save(evaluated, step3_path)
        status.update("step3_evaluate", "Evaluation complete", {
            "evaluated": len(candidates_to_eval),
            "passed": len(evaluated),
            "failed": len(candidates_to_eval) - len(evaluated),
        })

    # ------------------------------------------------------------------
    # STEP 3.5: Verify dataset sizes
    # ------------------------------------------------------------------
    step35_path = intermediate / "step3_5_verified.json"
    if _checkpoint_exists(step35_path, resume):
        log.info("Step 3.5: loading from checkpoint")
        step35_data = _load(step35_path)
        verified = step35_data["passed"]
    else:
        status.update("step3_5_verify", f"Verifying dataset sizes for {len(evaluated)} papers...")
        from .verify_size import verify_paper_sizes
        verified, rejected = verify_paper_sizes(evaluated)
        _save({"passed": verified, "rejected": rejected}, step35_path)
        status.update("step3_5_verify", "Size verification complete", {
            "passed": len(verified),
            "rejected_oversize": len(rejected),
        })

    # ------------------------------------------------------------------
    # STEP 4: Score and rank
    # ------------------------------------------------------------------
    status.update("step4_score", "Scoring and applying diversity adjustment...")
    from .score import score_and_rank
    final_papers = score_and_rank(verified, max_papers=max_papers)
    _save(final_papers, intermediate / "step4_scored.json")
    status.update("step4_score", "Scoring complete", {
        "papers_selected": len(final_papers),
        "domains": len({p.get("domain", "?") for p in final_papers}),
    })

    # ------------------------------------------------------------------
    # STEP 5: Write output JSON
    # ------------------------------------------------------------------
    # Strip internal fields before output
    output_papers = []
    for p in final_papers:
        clean = {k: v for k, v in p.items()
                 if k not in ("size_details", "size_verified",
                              "is_open_access", "pmid", "pmcid", "abstract",
                              "authors", "source", "include", "exclusion_reason")}
        output_papers.append(clean)

    output = {
        "metadata": {
            "generated_date": date.today().isoformat(),
            "date_window_start": start_date,
            "date_window_end": end_date,
            "total_candidates_screened": len(candidates),
            "total_papers_included": len(output_papers),
            "prompt_version": "v3_python",
        },
        "papers": output_papers,
    }

    output_file = run_dir / f"computational_biology_papers_{date.today().year}.json"
    _save(output, output_file)
    log.info("Wrote %d papers to %s", len(output_papers), output_file)

    # Verify JSON
    _load(output_file)  # will raise if invalid
    status.update("step5_output", "Output generated", {
        "papers_included": len(output_papers),
        "output_file": str(output_file),
    })

    # ------------------------------------------------------------------
    # Review gate — pause for user to inspect the output before downloading
    # ------------------------------------------------------------------
    if review:
        print(f"\n{'=' * 70}")
        print(f"  REVIEW: {len(output_papers)} papers selected")
        print(f"  Output: {output_file}")
        print(f"{'=' * 70}")
        for i, p in enumerate(output_papers, 1):
            print(f"\n  {i}. [{p.get('quality_score', '?')}] {p.get('title', '?')[:80]}")
            print(f"     {p.get('journal', '?')} | {p.get('domain', '?')} | {p.get('doi', '?')}")
            acc = [a.get('accession', '') for a in p.get('dataset_accession', [])]
            print(f"     Datasets: {', '.join(acc)} ({p.get('estimated_processed_data_size_gb', '?')} GB)")
        print(f"\n{'=' * 70}")
        print(f"  Edit {output_file} to remove unwanted papers, then continue.")
        print(f"  To resume: python -m kosmos_pipeline -o {output_dir}")
        print(f"{'=' * 70}\n")

        try:
            answer = input("Proceed with download and upload? [y/N] ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            answer = "n"

        if answer != "y":
            log.info("Stopped at review gate. Re-run to continue.")
            status.update("review", "Paused for review", {
                "papers_included": len(output_papers),
                "output_file": str(output_file),
            })
            return output

        # Reload output in case user edited the file
        reloaded = _load(output_file)
        final_papers = reloaded.get("papers", final_papers)
        log.info("Continuing with %d papers after review", len(final_papers))

    # ------------------------------------------------------------------
    # STEP 5.5: Download paper PDFs
    # ------------------------------------------------------------------
    status.update("step5_5_pdfs", f"Downloading PDFs for {len(final_papers)} papers...")
    from .download import download_pdfs
    pdfs_dir = run_dir / "pdfs"
    pdf_results = download_pdfs(final_papers, pdfs_dir)
    _save(pdf_results, intermediate / "step5_5_pdf_results.json")
    downloaded_count = sum(1 for r in pdf_results.values() if r["downloaded"])
    status.update("step5_5_pdfs", "PDF downloads complete", {
        "downloaded": downloaded_count,
        "failed": len(pdf_results) - downloaded_count,
    })

    # ------------------------------------------------------------------
    # STEP 6: Download datasets (optional)
    # ------------------------------------------------------------------
    if not skip_download:
        status.update("step6_download", f"Downloading datasets for {len(final_papers)} papers...")
        from .download import download_all
        datasets_dir = run_dir / "datasets"
        download_results = download_all(final_papers, datasets_dir)
        _save(download_results, intermediate / "step6_download_results.json")
        status.update("step6_download", "Downloads complete", {
            "papers_downloaded": len(download_results),
        })
    else:
        log.info("Skipping dataset download (--skip-download)")

    # ------------------------------------------------------------------
    # STEP 7: Upload to Edison (optional)
    # ------------------------------------------------------------------
    if not skip_upload and not skip_download:
        kopt_dir = Path(kosmos_opt_dir) if kosmos_opt_dir else None
        status.update("step7_upload", f"Uploading {len(final_papers)} datasets to Edison...")
        from .upload import upload_all, DEFAULT_KOSMOS_OPT_DIR
        probes_dir = run_dir / "probes"
        upload_results = upload_all(
            final_papers,
            datasets_dir=run_dir / "datasets",
            probes_dir=probes_dir,
            stage=stage,
            kosmos_opt_dir=kopt_dir or DEFAULT_KOSMOS_OPT_DIR,
        )
        _save(upload_results, intermediate / "step7_upload_results.json")
        status.update("step7_upload", "Uploads complete", upload_results)
    elif skip_upload:
        log.info("Skipping Edison upload (--skip-upload)")

    # ------------------------------------------------------------------
    # Manifest: single file linking paper → PDF → probe → data_entry
    # ------------------------------------------------------------------
    manifest = []
    for p in final_papers:
        doi = p.get("doi", "unknown")
        name = None
        for acc in p.get("dataset_accession", []):
            a = acc.get("accession", "")
            if a.upper().startswith("GSE"):
                name = a.upper()
                break
            if "zenodo" in a.lower():
                import re as _re
                m = _re.search(r'(\d{5,})', a)
                if m:
                    name = f"Zenodo_{m.group(1)}"
                    break
        if not name:
            import re as _re
            name = _re.sub(r'[^\w.-]', '_', doi)

        entry = {
            "doi": doi,
            "title": p.get("title", ""),
            "first_author": p.get("first_author", ""),
            "journal": p.get("journal", ""),
            "quality_score": p.get("quality_score", 0),
            "domain": p.get("domain", ""),
            "probe_name": name,
        }

        # PDF path
        pdf_info = pdf_results.get(doi, {})
        entry["pdf"] = pdf_info.get("path", "")

        # Upload / data_entry info
        upload_info = upload_results.get(doi, {})
        if isinstance(upload_info, dict):
            entry["data_entry"] = upload_info.get("data_entry", "")
            entry["probe_dir"] = upload_info.get("probe_dir", "")
        else:
            entry["data_entry"] = ""
            entry["probe_dir"] = ""

        # Dataset accessions
        entry["dataset_accession"] = [
            acc.get("accession", "") for acc in p.get("dataset_accession", [])
        ]

        manifest.append(entry)

    manifest_path = run_dir / "manifest.json"
    _save(manifest, manifest_path)
    log.info("Wrote manifest: %s", manifest_path)

    # ------------------------------------------------------------------
    # Done
    # ------------------------------------------------------------------
    status.update("done", "Pipeline complete", {
        "papers_included": len(output_papers),
        "distinct_domains": len({p.get("domain", "?") for p in output_papers}),
        "output_file": str(output_file),
        "manifest_file": str(manifest_path),
    })

    log.info("Pipeline complete: %d papers → %s", len(output_papers), output_file)
    return output


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv=None):
    p = argparse.ArgumentParser(
        prog="kosmos_pipeline",
        description="Kosmos benchmark paper search and curation pipeline",
    )
    p.add_argument(
        "--output-dir", "-o",
        default=f"./runs/{date.today().isoformat()}",
        help="Output directory for this run (default: ./runs/YYYY-MM-DD)",
    )
    p.add_argument(
        "--no-resume",
        action="store_true",
        help="Do not resume from checkpoints — start fresh",
    )
    p.add_argument(
        "--max-candidates",
        type=int,
        default=config.MAX_CANDIDATES_TO_EVALUATE,
        help=f"Max candidates to evaluate via Claude API (default: {config.MAX_CANDIDATES_TO_EVALUATE})",
    )
    p.add_argument(
        "--max-papers",
        type=int,
        default=config.MAX_PAPERS_OUTPUT,
        help=f"Max papers in final output (default: {config.MAX_PAPERS_OUTPUT})",
    )
    p.add_argument(
        "--model",
        default=config.DEFAULT_MODEL,
        help=f"Claude model for evaluation (default: {config.DEFAULT_MODEL})",
    )
    p.add_argument(
        "--exclude-dois",
        help="File with DOIs to exclude (one per line, or .xlsx)",
    )
    p.add_argument(
        "--review",
        action="store_true",
        help="Pause after paper selection to review before downloading/uploading",
    )
    p.add_argument(
        "--skip-download",
        action="store_true",
        help="Skip dataset download step",
    )
    p.add_argument(
        "--skip-upload",
        action="store_true",
        help="Skip Edison upload step",
    )
    p.add_argument(
        "--stage",
        default="DEV",
        choices=["DEV", "STAGING", "PROD"],
        help="Edison environment stage (default: DEV)",
    )
    p.add_argument(
        "--kosmos-opt-dir",
        help=f"Path to kosmos-opt repo (default: ~/kosmos-opt)",
    )
    p.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug logging",
    )
    return p.parse_args(argv)


def main(argv=None):
    args = parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s — %(message)s",
        datefmt="%H:%M:%S",
    )

    run_pipeline(
        output_dir=args.output_dir,
        resume=not args.no_resume,
        max_candidates=args.max_candidates,
        max_papers=args.max_papers,
        model=args.model,
        exclude_dois_file=args.exclude_dois,
        skip_download=args.skip_download,
        skip_upload=args.skip_upload,
        review=args.review,
        stage=args.stage,
        kosmos_opt_dir=args.kosmos_opt_dir,
    )
